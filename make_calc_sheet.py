# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter as L

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '세액공제계산'

# 스타일
navy = PatternFill('solid', fgColor='1a4fa8')
lgray = PatternFill('solid', fgColor='F2F2F2')
white_fill = PatternFill('solid', fgColor='FFFFFF')
hdr_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
title_font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
bold_font = Font(name='맑은 고딕', size=10, bold=True)
normal_font = Font(name='맑은 고딕', size=10)
small_font = Font(name='맑은 고딕', size=9, color='888888')
thin = Side(style='thin', color='000000')
bd = Border(top=thin, bottom=thin, left=thin, right=thin)
center = Alignment(horizontal='center', vertical='center', wrapText=True)
left_al = Alignment(horizontal='left', vertical='center')
right_al = Alignment(horizontal='right', vertical='center')
won_fmt = '#,##0'
dec_fmt = '#,##0.00'

def sc(r, c, val, font=normal_font, fill=white_fill, align=center, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.fill = fill
    cell.border = bd
    cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell

def hdr(r, c, val):
    return sc(r, c, val, font=hdr_font, fill=navy)

def fc(r, c, formula, fmt=won_fmt):
    cell = ws.cell(row=r, column=c)
    cell.value = formula
    cell.font = normal_font
    cell.fill = white_fill
    cell.border = bd
    cell.alignment = right_al
    cell.number_format = fmt
    return cell

# 열 너비
for c in range(1, 19):
    ws.column_dimensions[L(c)].width = 15
ws.column_dimensions['N'].width = 18  # 케이스 열 넓게

# ═══════════════════════════════════════
# 1. 상시근로자수 테이블
# ═══════════════════════════════════════
r = 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
sc(r, 1, '고용증대 세액공제 계산 (중소기업)', font=title_font, fill=navy, align=left_al)
for c in range(2, 8):
    sc(r, c, None, fill=navy)

r = 3
hdr(r, 1, '연도')
hdr(r, 2, '청년 수도권')
hdr(r, 3, '청년 지방')
hdr(r, 4, '비청년 수도권')
hdr(r, 5, '비청년 지방')
hdr(r, 6, '합계')

raw_data = {
    2017: (5.25, 1.75, 6.91, 6.16),
    2018: (7.00, 2.75, 6.33, 5.58),
    2019: (5.58, 1.41, 8.37, 7.00),
    2020: (7.00, 1.83, 8.66, 8.50),
    2021: (9.58, 1.75, 12.33, 11.00),
    2022: (8.16, 1.00, 10.08, 12.58),
}

data_start_row = 4
year_rows = {}
for i, (yr, vals) in enumerate(sorted(raw_data.items())):
    rr = data_start_row + i
    year_rows[yr] = rr
    sc(rr, 1, f'{yr}년', font=bold_font)
    for ci, v in enumerate(vals, 2):
        sc(rr, ci, v, fmt=dec_fmt, align=right_al)
    fc(rr, 6, f'=SUM(B{rr}:E{rr})', dec_fmt)

# ═══════════════════════════════════════
# 2. 단가표
# ═══════════════════════════════════════
r = 11
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
sc(r, 1, '단가 (만원) - 중소기업', font=bold_font, fill=lgray)
for c in range(2, 6):
    sc(r, c, None, fill=lgray)

r = 12
sc(r, 1, '구분', font=bold_font, fill=lgray)
sc(r, 2, '청년 수도권', font=bold_font, fill=lgray)
sc(r, 3, '청년 지방', font=bold_font, fill=lgray)
sc(r, 4, '비청년 수도권', font=bold_font, fill=lgray)
sc(r, 5, '비청년 지방', font=bold_font, fill=lgray)

r = 13
sc(r, 1, '기본단가')
sc(r, 2, 1100, fmt=won_fmt, align=right_al)
sc(r, 3, 1200, fmt=won_fmt, align=right_al)
sc(r, 4, 700, fmt=won_fmt, align=right_al)
sc(r, 5, 770, fmt=won_fmt, align=right_al)

r = 14
sc(r, 1, '21-22 지방청년 특례', font=small_font)
sc(r, 3, 1300, fmt=won_fmt, align=right_al, font=bold_font)

# 단가 셀 참조
REF_YM = '$B$13'
REF_YL = '$C$13'
REF_YL_HIST = '$C$14'
REF_NM = '$D$13'
REF_NL = '$E$13'

# ═══════════════════════════════════════
# 3. 시작연도별 계산
# ═══════════════════════════════════════
cur_row = 17
results = {}

for startY in range(2018, 2023):
    prevY = startY - 1

    # 단가 참조
    yl_ref = REF_YL_HIST if startY in [2021, 2022] else REF_YL

    # 제목
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=18)
    sc(cur_row, 1, f'{startY}년 시작 (전기={prevY}, 당기={startY})', font=title_font, fill=navy, align=left_al)
    for c in range(2, 19):
        sc(cur_row, c, None, fill=navy)
    cur_row += 1

    # 증가인원 (3단계 cap)
    # 1단계: 각 카테고리별 MAX(0, diff)
    # 2단계: 수도권 한도 = MAX(0, ym_diff + nm_diff), 청년 우선
    # 3단계: 비수도권 한도 = MAX(0, yl_diff + nl_diff), 청년 우선
    # 4단계: 전체 한도 = MAX(0, 전체diff), 2+3 합이 초과하면 비례축소
    hdr(cur_row, 1, '구분')
    hdr(cur_row, 2, '청년 수도권')
    hdr(cur_row, 3, '청년 지방')
    hdr(cur_row, 4, '비청년 수도권')
    hdr(cur_row, 5, '비청년 지방')
    hdr(cur_row, 6, '합계')
    cur_row += 1

    pr = year_rows[prevY]
    dr = year_rows[startY]

    # ① diff행 (음수 가능)
    diff_row = cur_row
    sc(cur_row, 1, '① diff', font=small_font)
    fc(cur_row, 2, f'=B{dr}-B{pr}', dec_fmt)
    fc(cur_row, 3, f'=C{dr}-C{pr}', dec_fmt)
    fc(cur_row, 4, f'=D{dr}-D{pr}', dec_fmt)
    fc(cur_row, 5, f'=E{dr}-E{pr}', dec_fmt)
    fc(cur_row, 6, f'=SUM(B{cur_row}:E{cur_row})', dec_fmt)
    cur_row += 1

    # ② 지역별 한도
    cap_row = cur_row
    sc(cur_row, 1, '② 지역한도', font=small_font)
    fc(cur_row, 2, f'=MAX(0,B{diff_row}+D{diff_row})', dec_fmt)  # 수도권cap
    fc(cur_row, 3, f'=MAX(0,C{diff_row}+E{diff_row})', dec_fmt)  # 비수도권cap
    fc(cur_row, 4, f'=MAX(0,F{diff_row})', dec_fmt)  # 전체cap
    sc(cur_row, 5, '', font=small_font)
    sc(cur_row, 6, '', font=small_font)
    cur_row += 1

    # ③ 지역cap 적용 (청년 우선)
    region_row = cur_row
    sc(cur_row, 1, '③ 지역cap후', font=small_font)
    fc(cur_row, 2, f'=MIN(MAX(0,B{diff_row}),B{cap_row})', dec_fmt)  # 청년수도
    fc(cur_row, 3, f'=MIN(MAX(0,C{diff_row}),C{cap_row})', dec_fmt)  # 청년지방
    fc(cur_row, 4, f'=MAX(0,MIN(MAX(0,D{diff_row}),B{cap_row}-B{cur_row}))', dec_fmt)  # 비청년수도
    fc(cur_row, 5, f'=MAX(0,MIN(MAX(0,E{diff_row}),C{cap_row}-C{cur_row}))', dec_fmt)  # 비청년지방
    fc(cur_row, 6, f'=SUM(B{cur_row}:E{cur_row})', dec_fmt)
    cur_row += 1

    # ④ 전체cap 적용 (공제단가 큰 순서로 배정)
    # 단가순: 청년지방(1200) > 청년수도(1100) > 비청년지방(770) > 비청년수도(700)
    # 2021-2022 특례: 청년지방 1300
    # 순서: C(청년지방) → B(청년수도) → E(비청년지방) → D(비청년수도)
    inc_row = cur_row
    rr = region_row
    tc = f'D{cap_row}'  # 전체cap 셀
    sc(cur_row, 1, '④ 최종증가', font=bold_font)
    # 1순위: 청년지방 = MIN(지역cap후, 전체cap)
    fc(cur_row, 3,
       f'=MIN(C{rr},{tc})',
       dec_fmt)
    # 2순위: 청년수도 = MIN(지역cap후, 전체cap - 1순위)
    fc(cur_row, 2,
       f'=MIN(B{rr},MAX(0,{tc}-C{cur_row}))',
       dec_fmt)
    # 3순위: 비청년지방 = MIN(지역cap후, 전체cap - 1순위 - 2순위)
    fc(cur_row, 5,
       f'=MIN(E{rr},MAX(0,{tc}-C{cur_row}-B{cur_row}))',
       dec_fmt)
    # 4순위: 비청년수도 = MIN(지역cap후, 전체cap - 1~3순위)
    fc(cur_row, 4,
       f'=MIN(D{rr},MAX(0,{tc}-C{cur_row}-B{cur_row}-E{cur_row}))',
       dec_fmt)
    fc(cur_row, 6, f'=SUM(B{cur_row}:E{cur_row})', dec_fmt)
    cur_row += 2

    # 연도별 공제/추징
    # A=과세연도 B=구분 C~F=증가인원 G=공제액
    # H=당기합계 I=해당연도합계 J=전체감소 K=당기청년 L=해당청년 M=청년감소
    # N=케이스 O=추징산출 P=추징한도 Q=추징액 R=순공제
    hdr(cur_row, 1, '과세연도')
    hdr(cur_row, 2, '구분')
    hdr(cur_row, 3, '증가(청년수도)')
    hdr(cur_row, 4, '증가(청년지방)')
    hdr(cur_row, 5, '증가(비청년수도)')
    hdr(cur_row, 6, '증가(비청년지방)')
    hdr(cur_row, 7, '공제액(원)')
    hdr(cur_row, 8, '당기합계')
    hdr(cur_row, 9, '해당연도합계')
    hdr(cur_row, 10, '전체감소')
    hdr(cur_row, 11, '당기청년합계')
    hdr(cur_row, 12, '해당청년합계')
    hdr(cur_row, 13, '청년감소')
    hdr(cur_row, 14, '케이스')
    hdr(cur_row, 15, '추징산출(원)')
    hdr(cur_row, 16, '추징한도(원)')
    hdr(cur_row, 17, '추징액(원)')
    hdr(cur_row, 18, '순공제액(원)')
    cur_row += 1

    results[startY] = {}
    danggi_r = year_rows[startY]
    prev_r = year_rows[prevY]

    # 타임라인 구성
    timeline = [('danggi', startY)]
    for i in range(1, 3):
        cy = startY + i
        if cy <= 2022:
            timeline.append((f'{i}차추가', cy))

    first_data_row = cur_row
    for idx, (role, cy) in enumerate(timeline):
        role_label = '당기' if role == 'danggi' else role
        sc(cur_row, 1, f'{cy}년', font=bold_font)
        sc(cur_row, 2, role_label)

        # C~F: 증가인원
        for ci in range(3, 7):
            real_ci = ci - 1
            fc(cur_row, ci, f'={L(real_ci)}{inc_row}', dec_fmt)

        # G: 공제액
        fc(cur_row, 7,
           f'=TRUNC(C{cur_row}*{REF_YM}*10000)'
           f'+TRUNC(D{cur_row}*{yl_ref}*10000)'
           f'+TRUNC(E{cur_row}*{REF_NM}*10000)'
           f'+TRUNC(F{cur_row}*{REF_NL}*10000)',
           won_fmt)

        # H: 당기 전체합계
        fc(cur_row, 8, f'=F{danggi_r}', dec_fmt)

        if role == 'danggi':
            # 당기는 사후관리 없음
            if cy in year_rows:
                fc(cur_row, 9, f'=F{year_rows[cy]}', dec_fmt)
            sc(cur_row, 10, '', font=small_font)
            fc(cur_row, 11, f'=B{danggi_r}+C{danggi_r}', dec_fmt)  # 당기청년
            if cy in year_rows:
                fc(cur_row, 12, f'=B{year_rows[cy]}+C{year_rows[cy]}', dec_fmt)
            sc(cur_row, 13, '', font=small_font)
            sc(cur_row, 14, '-', font=small_font)
            sc(cur_row, 15, 0, fmt=won_fmt, align=right_al)
            sc(cur_row, 16, 0, fmt=won_fmt, align=right_al)
            sc(cur_row, 17, 0, fmt=won_fmt, align=right_al)
        else:
            if cy in year_rows:
                cy_r = year_rows[cy]
                mult = idx  # 1차=1, 2차=2

                # I: 해당연도 전체합계
                fc(cur_row, 9, f'=F{cy_r}', dec_fmt)
                # J: 전체감소 = MAX(0, 당기 - 해당연도)
                fc(cur_row, 10, f'=MAX(0,H{cur_row}-I{cur_row})', dec_fmt)
                # K: 당기 청년합계 = B+C (청년수도+청년지방)
                fc(cur_row, 11, f'=B{danggi_r}+C{danggi_r}', dec_fmt)
                # L: 해당연도 청년합계
                fc(cur_row, 12, f'=B{cy_r}+C{cy_r}', dec_fmt)
                # M: 청년감소 = MAX(0, 당기청년 - 해당청년)
                fc(cur_row, 13, f'=MAX(0,K{cur_row}-L{cur_row})', dec_fmt)

                # N: 케이스 판정
                # ok: 전체감소=0 AND 청년감소=0
                # youth_only(나.): 전체감소=0 AND 청년감소>0
                # case_a(가.㉠): 전체감소>0 AND 청년감소>=전체감소
                # case_b(가.㉡): 전체감소>0 AND 청년감소<전체감소
                fc(cur_row, 14,
                   f'=IF(AND(J{cur_row}=0,M{cur_row}=0),"OK",'
                   f'IF(AND(J{cur_row}=0,M{cur_row}>0),"나.청년감소",'
                   f'IF(M{cur_row}>=J{cur_row},"가.㉠전체(청년주)","가.㉡전체(비청년주)")))',
                   '@')

                # O: 추징산출
                # 나.youth_only: MIN(청년감소, 청년증가) × (청년단가-비청년단가) × mult
                # 가.㉠case_a: 전환분(MIN(청년감소,청년증가)-전체감소)×차액 + 이탈분(전체감소)×청년단가 × mult
                # 가.㉡case_b: 각 카테고리 감소 × 해당단가 × mult
                # 중소기업(수도권/지방 구분):
                youth_inc_ref = f'(C{first_data_row}+D{first_data_row})'  # 당기 증가 청년 합계

                fc(cur_row, 15,
                   f'=IF(N{cur_row}="OK",0,'
                   # 나. youth_only: 청년감소 × (청년단가-비청년단가) 수도권가중평균
                   f'IF(LEFT(N{cur_row},2)="나.",'
                   f'TRUNC(MIN(M{cur_row},{youth_inc_ref})*({REF_YM}-{REF_NM})*10000)*{mult},'
                   # 가.㉡ case_b: 각 카테고리별 감소 × 단가
                   f'IF(LEFT(N{cur_row},4)="가.㉡",'
                   f'(TRUNC(MAX(0,B{danggi_r}-B{cy_r})*{REF_YM}*10000)'
                   f'+TRUNC(MAX(0,C{danggi_r}-C{cy_r})*{yl_ref}*10000)'
                   f'+TRUNC(MAX(0,D{danggi_r}-D{cy_r})*{REF_NM}*10000)'
                   f'+TRUNC(MAX(0,E{danggi_r}-E{cy_r})*{REF_NL}*10000))'
                   f'*{mult},'
                   # 가.㉠ case_a: 전환분×차액 + 이탈분×청년단가
                   f'(TRUNC(MAX(0,MIN(M{cur_row},{youth_inc_ref})-J{cur_row})*({REF_YM}-{REF_NM})*10000)'
                   f'+TRUNC(J{cur_row}*{REF_YM}*10000))'
                   f'*{mult}'
                   f')))',
                   won_fmt)

                # P: 추징한도
                if idx == 1:
                    fc(cur_row, 16, f'=G{first_data_row}', won_fmt)
                else:
                    fc(cur_row, 16,
                       f'=G{first_data_row}+G{cur_row-1}-Q{cur_row-1}',
                       won_fmt)

                # Q: 추징액 = MIN(추징산출, 추징한도) - 이전추징 차감
                if idx == 1:
                    fc(cur_row, 17, f'=MIN(O{cur_row},P{cur_row})', won_fmt)
                else:
                    fc(cur_row, 17,
                       f'=MIN(MAX(0,O{cur_row}-Q{cur_row-1}),MAX(0,P{cur_row}))',
                       won_fmt)
            else:
                sc(cur_row, 9, '', font=small_font)
                sc(cur_row, 10, '', font=small_font)
                fc(cur_row, 11, f'=B{danggi_r}+C{danggi_r}', dec_fmt)
                sc(cur_row, 12, '', font=small_font)
                sc(cur_row, 13, '', font=small_font)
                sc(cur_row, 14, '데이터없음', font=small_font)
                sc(cur_row, 15, 0, fmt=won_fmt, align=right_al)
                sc(cur_row, 16, 0, fmt=won_fmt, align=right_al)
                sc(cur_row, 17, 0, fmt=won_fmt, align=right_al)

        # R: 순공제 = 공제 - 추징
        fc(cur_row, 18, f'=G{cur_row}-Q{cur_row}', won_fmt)

        results[startY][cy] = cur_row
        cur_row += 1

    # 합계행
    last_data_row = cur_row - 1
    sc(cur_row, 1, '합계', font=bold_font, fill=lgray)
    for c in range(2, 15):
        sc(cur_row, c, None, fill=lgray)
    for c in [7, 17, 18]:
        fc(cur_row, c, f'=SUM({L(c)}{first_data_row}:{L(c)}{last_data_row})', won_fmt)
        ws.cell(row=cur_row, column=c).fill = lgray
        ws.cell(row=cur_row, column=c).font = bold_font
    for c in [15, 16]:
        sc(cur_row, c, None, fill=lgray)

    results[startY]['sum'] = cur_row
    cur_row += 2

# ═══════════════════════════════════════
# 4. 종합 매트릭스
# ═══════════════════════════════════════
cal_years = sorted(set(cy for sy in results for cy in results[sy] if cy != 'sum'))

ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(cal_years)+2)
sc(cur_row, 1, '종합 매트릭스 (과세연도별 순공제액, 원)', font=title_font, fill=navy, align=left_al)
for c in range(2, len(cal_years)+3):
    sc(cur_row, c, None, fill=navy)
cur_row += 1

hdr(cur_row, 1, '시작 \\ 과세연도')
for i, cy in enumerate(cal_years):
    hdr(cur_row, i+2, f'{cy}년')
hdr(cur_row, len(cal_years)+2, '합계')
cur_row += 1

matrix_first = cur_row
for sy in sorted(results.keys()):
    sc(cur_row, 1, f'{sy}년 시작', font=bold_font)
    for i, cy in enumerate(cal_years):
        if cy in results[sy]:
            ref_row = results[sy][cy]
            fc(cur_row, i+2, f'=R{ref_row}', won_fmt)
        else:
            sc(cur_row, i+2, None)
    first_col = L(2)
    last_col = L(len(cal_years)+1)
    fc(cur_row, len(cal_years)+2, f'=SUM({first_col}{cur_row}:{last_col}{cur_row})', won_fmt)
    ws.cell(row=cur_row, column=len(cal_years)+2).font = bold_font
    cur_row += 1
matrix_last = cur_row - 1

# 연도별 합계
sc(cur_row, 1, '과세연도 합계', font=bold_font, fill=lgray)
for i in range(len(cal_years)+1):
    col = i + 2
    fc(cur_row, col, f'=SUM({L(col)}{matrix_first}:{L(col)}{matrix_last})', won_fmt)
    ws.cell(row=cur_row, column=col).fill = lgray
    ws.cell(row=cur_row, column=col).font = bold_font

wb.save('C:/Users/deoks/goyong-tonghap/고용증대_세액공제계산.xlsx')
print('OK')
